#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TGSimilarSpam - Telegram Channel Discovery & BD Messaging Bot
Recursively finds similar channels, identifies owners, sends bulk messages
"""

import asyncio
import re
import sys
import io
import json
import os
from datetime import datetime, timedelta
from collections import deque

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from telethon import TelegramClient
from telethon.tl.functions.channels import GetFullChannelRequest, GetChannelRecommendationsRequest
from telethon.errors import FloodWaitError
import openpyxl
from openpyxl.styles import PatternFill, Font

# ==================== CONFIG ====================
CONFIG_FILE = 'config.json'
SENT_LOG_FILE = 'sent_log.json'
DATA_FILE = 'contacts.xlsx'

# Load config
if not os.path.exists(CONFIG_FILE):
    print("[ERROR] config.json not found. Run setup.py first!")
    sys.exit(1)

with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
    CONFIG = json.load(f)

# Use credentials from config, or fallback to Telethon defaults
API_ID = CONFIG.get('api_id', 20190360)
API_HASH = CONFIG.get('api_hash', '67029a9453eb8a1f64fcead2fb0195b3')
PHONE = CONFIG.get('phone')
SEED_CHANNELS = CONFIG.get('seed_channels', [])
KEYWORDS = CONFIG.get('keywords', [])

COOLDOWN_DAYS = CONFIG.get('cooldown_days', 2)
SEND_DELAY_SECONDS = CONFIG.get('send_delay_seconds', 15)
MAX_SENT_PER_RUN = CONFIG.get('max_sent_per_run', 50)
TARGET_LANGUAGE = CONFIG.get('target_language', 'BOTH')

MSG_RUS = CONFIG.get('msg_ru', '')
MSG_ENG = CONFIG.get('msg_en', '')

if not MSG_RUS and not MSG_ENG:
    print("[ERROR] No messages in config.json")
    sys.exit(1)

# ==================== BOT ====================

class TelegramBDBot:
    def __init__(self):
        self.client = TelegramClient(f'session_{PHONE}', API_ID, API_HASH)
        self.processed_channels = set()
        self.queue = deque()
        self.results = []
        self.sent_count = 0
        self.error_count = 0
        self.sent_log = self.load_sent_log()
    
    # ==================== LOGGING ====================
    
    def load_sent_log(self):
        """Load sent log from JSON"""
        if os.path.exists(SENT_LOG_FILE):
            try:
                with open(SENT_LOG_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def save_sent_log(self):
        """Save sent log to JSON"""
        with open(SENT_LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.sent_log, f, ensure_ascii=False, indent=2)
    
    def is_owner_in_cooldown(self, owner):
        """Check if owner is in cooldown period"""
        if owner not in self.sent_log:
            return False
        
        last_sent = datetime.fromisoformat(self.sent_log[owner])
        days_passed = (datetime.now() - last_sent).days
        
        if days_passed < COOLDOWN_DAYS:
            return True
        return False
    
    def mark_as_sent(self, owner):
        """Mark owner as sent"""
        self.sent_log[owner] = datetime.now().isoformat()
        self.save_sent_log()
    
    # ==================== LANGUAGE DETECTION ====================
    
    def is_cyrillic_char(self, char):
        """Check if character is Cyrillic"""
        return 0x0400 <= ord(char) <= 0x04FF
    
    def detect_language_from_about(self, about):
        """Detect language from channel description"""
        if not about:
            return None
        
        cyrillic = sum(1 for c in about if self.is_cyrillic_char(c))
        latin = sum(1 for c in about if c.isalpha() and not self.is_cyrillic_char(c))
        
        total = cyrillic + latin
        if total == 0:
            return None
        
        cyrillic_pct = (cyrillic / total) * 100
        
        if cyrillic_pct >= 30:
            return 'RU'
        elif cyrillic_pct < 10:
            return 'EN'
        else:
            return None
    
    async def detect_language_from_posts(self, channel_name):
        """Detect language from last posts"""
        try:
            messages = await self.client.get_messages(channel_name, limit=5)
            
            cyrillic = 0
            latin = 0
            
            for msg in messages:
                if msg.text:
                    for char in msg.text:
                        if char.isalpha():
                            if self.is_cyrillic_char(char):
                                cyrillic += 1
                            else:
                                latin += 1
            
            total = cyrillic + latin
            if total == 0:
                return None
            
            cyrillic_pct = (cyrillic / total) * 100
            
            if cyrillic_pct >= 30:
                return 'RU'
            elif cyrillic_pct < 10:
                return 'EN'
            else:
                return None
        
        except Exception:
            return None
    
    # ==================== SEARCH ====================
    
    async def get_owners(self, channel_name):
        """Get channel owners and language"""
        try:
            ent = await self.client.get_entity(channel_name)
            
            try:
                full = await self.client(GetFullChannelRequest(ent))
                about = full.full_chat.about or ''
            except:
                about = ''
            
            # Language detection
            lang = await self.detect_language_from_posts(channel_name)
            if lang is None:
                about_lang = self.detect_language_from_about(about)
                if about_lang is not None:
                    lang = about_lang
            if lang is None:
                lang = 'EN'
            
            # Extract usernames from about
            usernames = re.findall(r'@[\w_]+', about)
            owners = [u.lstrip('@').lower() for u in usernames if u.lstrip('@').lower() != channel_name.lower()]
            owners = list(set(owners))
            
            if owners:
                return owners[0], lang
            return None, lang
        
        except Exception:
            return None, 'EN'
    
    async def get_similar_channels(self, channel_name):
        """Get similar channels via Telegram API"""
        try:
            ent = await self.client.get_entity(channel_name)
            result = await self.client(GetChannelRecommendationsRequest(channel=ent))
            
            similar = []
            for ch in result.chats:
                username = getattr(ch, 'username', None)
                if username and username not in self.processed_channels:
                    similar.append(username)
            
            return similar[:10]
        
        except FloodWaitError as e:
            await asyncio.sleep(min(e.seconds + 1, 10))
            return []
        except Exception:
            return []
    
    # ==================== HISTORY CHECK ====================
    
    async def check_history_with_owner(self, owner):
        """Check conversation history with owner"""
        try:
            user = await self.client.get_entity(owner)
            messages = await self.client.get_messages(user, limit=30)
            
            my_messages_to_delete = []
            
            if messages:
                last_msg = messages[0]
                if not last_msg.out:
                    return False, []
            
            for msg in messages:
                if msg.out:
                    my_messages_to_delete.append(msg)
            
            if my_messages_to_delete:
                return True, my_messages_to_delete
            else:
                return True, []
        
        except Exception:
            return True, []
    
    # ==================== SENDING ====================
    
    async def send_message(self, owner, lang):
        """Send message to owner"""
        try:
            # Language filter
            if TARGET_LANGUAGE != 'BOTH':
                if lang != TARGET_LANGUAGE:
                    return f'lang_skip_{lang}'
            
            # Bot filter - skip accounts ending with "bot"
            if owner.lower().endswith('bot'):
                return 'bot_skipped'
            
            # Cooldown check
            if self.is_owner_in_cooldown(owner):
                return 'cooldown'
            
            should_send, old_messages = await self.check_history_with_owner(owner)
            
            if not should_send:
                return 'skipped'
            
            # Delete old messages
            if old_messages:
                try:
                    user = await self.client.get_entity(owner)
                    for old_msg in old_messages:
                        try:
                            await self.client.delete_messages(user, old_msg)
                        except:
                            try:
                                await self.client.delete_messages(user, [old_msg.id])
                            except:
                                pass
                    await asyncio.sleep(0.5)
                except Exception:
                    pass
            
            # Send message
            message = MSG_RUS if lang == 'RU' else MSG_ENG
            user = await self.client.get_entity(owner)
            
            await asyncio.wait_for(
                self.client.send_message(user, message, link_preview=False),
                timeout=15
            )
            
            self.mark_as_sent(owner)
            self.sent_count += 1
            
            await asyncio.sleep(SEND_DELAY_SECONDS)
            
            return 'sent'
        
        except asyncio.TimeoutError:
            self.error_count += 1
            return 'timeout'
        except Exception:
            self.error_count += 1
            return 'error'
    
    # ==================== PROCESSING ====================
    
    async def process_channel(self, channel_name, depth=0, send=True, max_sent=50):
        """Process channel"""
        if self.sent_count >= max_sent:
            return
        
        if channel_name in self.processed_channels or depth > 2:
            return
        
        self.processed_channels.add(channel_name)
        
        indent = '  ' * depth
        print(f"\n{indent}[L{depth}] {channel_name}...", end=" ")
        
        owner, lang = await self.get_owners(channel_name)
        
        if owner:
            print(f"FOUND {lang} | @{owner}", end=" ")
            
            if send:
                status = await self.send_message(owner, lang)
                status_label = {
                    'sent': '[SENT]',
                    'skipped': '[SKIPPED]',
                    'cooldown': '[COOLDOWN]',
                    'bot_skipped': '[BOT]',
                    'lang_skip_RU': '[LANG_SKIP_RU]',
                    'lang_skip_EN': '[LANG_SKIP_EN]',
                    'timeout': '[TIMEOUT]',
                }
                print(f"-> {status_label.get(status, status)}")
            else:
                print("[TEST]")
                status = 'test'
            
            result = {
                'channel': channel_name,
                'owner': owner,
                'lang': lang,
                'status': status if send else 'test',
                'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            self.results.append(result)
            # Save ALL results - both successful and failed
            self.save_result_incremental(result)
        else:
            print(f"NOT_FOUND {lang}")
            # Save "no owner found" as well
            result = {
                'channel': channel_name,
                'owner': 'NOT_FOUND',
                'lang': lang,
                'status': 'no_owner',
                'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            self.results.append(result)
            self.save_result_incremental(result)
        
        # Recursion
        if depth <= 1:
            print(f"{indent}  -> Looking for similar...", end=" ")
            try:
                similar = await self.get_similar_channels(channel_name)
                if similar:
                    print(f"found {len(similar)}")
                    for sim in similar:
                        if sim not in self.processed_channels and self.sent_count < max_sent:
                            self.queue.append((sim, depth + 1))
                else:
                    print("none")
            except Exception:
                print("error")
        
        await asyncio.sleep(1)
    
    # ==================== SAVING ====================
    
    def save_result_incremental(self, result):
        """Save result to Excel (incremental - adds rows, doesn't overwrite)"""
        try:
            if os.path.exists(DATA_FILE):
                # File exists - just add new row
                wb = openpyxl.load_workbook(DATA_FILE)
                ws = wb.active
                last_row = ws.max_row + 1
            else:
                # Create new file with headers
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # Only 2 columns: Channel Link and Owner Username
                headers = ['Channel Link', 'Owner Username']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(1, col, h)
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
                
                last_row = 2
            
            # Add data (only 2 columns)
            ws[f'A{last_row}'] = f"https://t.me/{result['channel']}"
            ws[f'B{last_row}'] = f"@{result['owner']}"
            
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 25
            
            wb.save(DATA_FILE)
        except Exception:
            pass
    
    # ==================== MAIN LOOP ====================
    
    async def run(self, send=True, max_sent=50):
        """Main loop"""
        for ch in SEED_CHANNELS:
            self.queue.append((ch, 0))
        
        while self.queue and self.sent_count < max_sent:
            channel_name, depth = self.queue.popleft()
            
            if channel_name not in self.processed_channels:
                await self.process_channel(channel_name, depth, send=send, max_sent=max_sent)
    
    async def close(self):
        await self.client.disconnect()


# ==================== MAIN ====================

async def main():
    bot = TelegramBDBot()
    
    try:
        print("[*] Connecting...")
        await bot.client.connect()
        
        if not await bot.client.is_user_authorized():
            print("[ERROR] Not authorized. Run setup.py first!")
            await bot.client.disconnect()
            sys.exit(1)
        
        me = await bot.client.get_me()
        print(f"[OK] {me.first_name}\n")
        
        print(f"[*] MODE: SEND")
        print(f"    Seed channels: {len(SEED_CHANNELS)}")
        print(f"    Keywords: {', '.join(KEYWORDS[:3])}...")
        print(f"    Target language: {TARGET_LANGUAGE}")
        print(f"    Max messages: {MAX_SENT_PER_RUN}\n")
        
        await bot.run(send=True, max_sent=MAX_SENT_PER_RUN)
        
        print(f"\n{'='*70}")
        print(f"[RESULTS]")
        print(f"  Contacts found: {len(bot.results)}")
        print(f"  Messages sent: {bot.sent_count}")
        print(f"  Errors: {bot.error_count}")
        print(f"{'='*70}\n")

    except Exception as e:
        print(f"[ERROR] {e}")
    
    finally:
        await bot.close()


if __name__ == '__main__':
    asyncio.run(main())
